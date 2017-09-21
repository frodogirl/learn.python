"""
Задача №2. Вывод в Эксель
-------------------------

Файл: `excel.py`

Ого! Оказывается в большом городе у людей много друзей. Чтобы хоть как-то
их систематизировать Хельга решила сохранить информацию о них в Excel.

Сделать так, чтобы результат работы предыдущей задачи выводился не 
только в консоль, но и в Эксель файл в три колонки: имя, 
фамилия, ссылка на профиль Вконтакте.
"""

from friedns import get_friends_list
from openpyxl import Workbook
from openpyxl.styles import Font

def set_column_title(ws, column, title):
    cell = ws.cell(row=1, column=column)
    cell.value = title
    cell.font = Font(bold=True)
    return ws


def ouptut_to_excel(friends):
    # TODO: сохранять список друзей в .xlsx файл с помощью библиотеки openpyxl
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = 'friends'

    work_sheet = set_column_title(work_sheet, 1, 'id')
    work_sheet = set_column_title(work_sheet, 2, 'first name')
    work_sheet = set_column_title(work_sheet, 3, 'last name')
    
    row = 2
    for friend in friends:
        work_sheet.cell(row=row, column=1).value = 'id' + str(friend.get('uid', 'n/a'))
        work_sheet.cell(row=row, column=2).value = friend.get('first_name', 'n/a')
        work_sheet.cell(row=row, column=3).value = friend.get('last_name', 'n/a')
        row += 1

    work_book.save(filename='friends.xlsx')

if __name__ == '__main__':
    friends = get_friends_list()
    ouptut_to_excel(friends)
    print('ok')
